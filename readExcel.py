import sys
sys.path.append("ExcelLib\\jdcal-1.2")
sys.path.append("ExcelLib\\et_xmlfile-1.0.1")
sys.path.append("ExcelLib\\openpyxl-2.3.3")
import openpyxl

wb = openpyxl.load_workbook('example.xlsx')
activeSheet = wb.active
print(activeSheet['B1'].value)
print(activeSheet.cell(row=2,column=2).value)
maxrows = activeSheet.max_row
maxcols = activeSheet.max_column
print(maxrows,maxcols)

for r in range(1,maxrows+1):
	for c in range(1,maxcols+1):
		print(activeSheet.cell(row=r,column=c).value)

sheet2 = wb.get_sheet_by_name('Sheet2')
maxrows = sheet2.max_row
maxcols = sheet2.max_column
print(maxrows,maxcols)

for r in range(1,maxrows+1):
	for c in range(1,maxcols+1):
		print(activeSheet.cell(row=r,column=c).value)